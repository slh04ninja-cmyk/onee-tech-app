import streamlit as st
import math
import io
from datetime import datetime

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="ONEE Tech Assistant",
    page_icon="⚡",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# ── Session state init ─────────────────────────────────────────────────────────
for key in ["res_t1", "res_t2"]:
    if key not in st.session_state:
        st.session_state[key] = None

# ── Custom CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Barlow:wght@400;600;700;800&family=Barlow+Condensed:wght@700;800&display=swap');

:root {
    --onee-green:  #00793B;
    --onee-dark:   #00501F;
    --onee-light:  #E8F5EE;
    --onee-accent: #F4A800;
    --onee-red:    #D32F2F;
    --bg:          #F4F6F0;
    --card:        #FFFFFF;
    --text:        #1A2E1A;
    --muted:       #5A7A5A;
    --border:      #C8DCC8;
}

* { font-family: 'Barlow', sans-serif; }

.stApp {
    background: var(--bg);
    background-image:
        radial-gradient(circle at 10% 20%, rgba(0,121,59,0.07) 0%, transparent 50%),
        radial-gradient(circle at 90% 80%, rgba(244,168,0,0.06) 0%, transparent 50%);
}

/* ── Header ── */
.onee-header {
    background: linear-gradient(135deg, var(--onee-dark) 0%, var(--onee-green) 100%);
    border-radius: 20px;
    padding: 32px 36px 28px;
    margin-bottom: 28px;
    position: relative;
    overflow: hidden;
    box-shadow: 0 8px 32px rgba(0,80,31,0.25);
}
.onee-header::before {
    content: 'lightning';
    position: absolute;
    right: 28px; top: 50%;
    transform: translateY(-50%);
    font-size: 80px;
    opacity: 0.12;
    line-height: 1;
}
.onee-header::after {
    content: '';
    position: absolute;
    bottom: 0; left: 0; right: 0;
    height: 4px;
    background: linear-gradient(90deg, var(--onee-accent), transparent);
}
.onee-logo {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 13px;
    font-weight: 700;
    letter-spacing: 4px;
    color: rgba(255,255,255,0.6);
    text-transform: uppercase;
    margin-bottom: 6px;
}
.onee-title {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 34px;
    font-weight: 800;
    color: #FFFFFF;
    line-height: 1.1;
    margin: 0;
}
.onee-subtitle {
    font-size: 14px;
    color: rgba(255,255,255,0.65);
    margin-top: 8px;
    font-weight: 400;
}

/* ── Cards ── */
.calc-card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 16px;
    padding: 28px;
    margin-bottom: 20px;
    box-shadow: 0 2px 12px rgba(0,80,31,0.07);
}
.card-title {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 20px;
    font-weight: 700;
    color: var(--onee-dark);
    margin-bottom: 20px;
    padding-bottom: 12px;
    border-bottom: 2px solid var(--onee-light);
}

/* ── Result boxes ── */
.result-box {
    border-radius: 12px;
    padding: 20px 24px;
    margin-top: 20px;
    border-left: 5px solid;
}
.result-ok   { background:#E8F5EE; border-color:var(--onee-green); }
.result-warn { background:#FFF8E1; border-color:var(--onee-accent); }
.result-err  { background:#FFEBEE; border-color:var(--onee-red);   }

.result-label {
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 2px;
    text-transform: uppercase;
    margin-bottom: 4px;
    color: #333333 !important;
}
.result-value {
    font-family: 'Barlow Condensed', sans-serif;
    font-size: 40px;
    font-weight: 800;
    line-height: 1;
    margin-bottom: 6px;
}
.result-ok   .result-value { color: #00501F !important; }
.result-warn .result-value { color: #5C3D00 !important; }
.result-err  .result-value { color: #B71C1C !important; }
.result-msg {
    font-size: 14px;
    font-weight: 600;
    margin-top: 8px;
    color: #1A1A1A !important;
}

/* ── Force metric text BLACK ── */
[data-testid="metric-container"] {
    background: var(--onee-light);
    border-radius: 10px;
    padding: 12px 16px !important;
    border: 1px solid var(--border);
}
[data-testid="stMetricLabel"] p,
[data-testid="stMetricLabel"] div {
    color: #333333 !important;
    font-weight: 700 !important;
    font-size: 12px !important;
}
[data-testid="stMetricValue"] div {
    color: #00501F !important;
    font-weight: 800 !important;
    font-size: 22px !important;
}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
    background: var(--onee-light);
    border-radius: 12px;
    padding: 4px;
    gap: 4px;
    border: none;
    margin-bottom: 20px;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 8px;
    font-weight: 700;
    font-size: 14px;
    color: var(--muted);
    padding: 10px 20px;
    border: none;
    background: transparent;
}
.stTabs [aria-selected="true"] {
    background: var(--onee-green) !important;
    color: white !important;
}
.stTabs [data-baseweb="tab-border"] { display: none; }

/* ── Inputs labels BLACK ── */
.stNumberInput label, .stSlider label,
.stSelectbox label, .stRadio label {
    font-weight: 700 !important;
    font-size: 13px !important;
    color: #1A2E1A !important;
}
/* Force input background WHITE and text BLACK */
.stNumberInput input,
.stNumberInput > div > div > input,
div[data-baseweb="input"] input,
div[data-baseweb="base-input"] input {
    background-color: #FFFFFF !important;
    background: #FFFFFF !important;
    color: #1A2E1A !important;
    border-radius: 8px !important;
    border-color: var(--border) !important;
    font-weight: 600 !important;
    font-size: 15px !important;
    -webkit-text-fill-color: #1A2E1A !important;
}
div[data-baseweb="input"],
div[data-baseweb="base-input"] {
    background-color: #FFFFFF !important;
    background: #FFFFFF !important;
    border-radius: 8px !important;
}
/* Stepper buttons (+/-) */
.stNumberInput button {
    background-color: var(--onee-light) !important;
    color: var(--onee-dark) !important;
    border-color: var(--border) !important;
}

/* ── Calculate Button ── */
.stButton > button {
    background: linear-gradient(135deg, var(--onee-green), var(--onee-dark)) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-size: 15px !important;
    padding: 12px 32px !important;
    width: 100%;
    box-shadow: 0 4px 14px rgba(0,121,59,0.3) !important;
    transition: all 0.2s !important;
}
.stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 20px rgba(0,121,59,0.4) !important;
}

/* ── Download Buttons ── */
.dl-excel .stDownloadButton > button {
    background: white !important;
    color: var(--onee-green) !important;
    border: 2px solid var(--onee-green) !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-size: 14px !important;
    padding: 10px 24px !important;
    width: 100%;
}
.dl-pdf .stDownloadButton > button {
    background: white !important;
    color: #C62828 !important;
    border: 2px solid #C62828 !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-size: 14px !important;
    padding: 10px 24px !important;
    width: 100%;
}

/* ── Misc ── */
hr { border-color: var(--border) !important; margin: 20px 0 !important; }
.onee-footer {
    text-align: center;
    padding: 20px;
    color: var(--muted);
    font-size: 12px;
    margin-top: 32px;
    border-top: 1px solid var(--border);
}
.badge {
    display: inline-block;
    background: var(--onee-light);
    color: var(--onee-dark);
    font-size: 11px;
    font-weight: 700;
    padding: 3px 10px;
    border-radius: 20px;
    border: 1px solid var(--border);
}
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 2rem; }
</style>
""", unsafe_allow_html=True)

# ── Helper : générer Excel ─────────────────────────────────────────────────────
def make_excel(title, data_rows):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:30]
    gfill = PatternFill("solid", fgColor="00793B")
    lfill = PatternFill("solid", fgColor="E8F5EE")
    hfill = PatternFill("solid", fgColor="C8DCC8")
    thin  = Side(style="thin", color="C8DCC8")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    def wc(r, c, v, fill=None, font=None, align="left"):
        x = ws.cell(row=r, column=c, value=v)
        if fill: x.fill = fill
        if font: x.font = font
        x.alignment = Alignment(horizontal=align, vertical="center")
        x.border = bdr
    ws.merge_cells("A1:C1")
    wc(1,1,f"ONEE — {title}", fill=gfill,
       font=Font(bold=True,color="FFFFFF",size=14), align="center")
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A2:C2")
    wc(2,1,f"Genere le {datetime.now().strftime('%d/%m/%Y %H:%M')}",
       font=Font(italic=True,color="5A7A5A"), align="center")
    for c, t in enumerate(["Parametre","Valeur","Unite"],1):
        wc(3,c,t, fill=hfill, font=Font(bold=True,color="00501F"), align="center")
    for r,(p,v,u) in enumerate(data_rows,4):
        f = lfill if r%2==0 else None
        wc(r,1,p, fill=f, font=Font(bold=True))
        wc(r,2,v, fill=f, align="center")
        wc(r,3,u, fill=f, align="center")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf

# ── Helper : générer PDF ───────────────────────────────────────────────────────
def make_pdf(title, data_rows, statut):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    ONEE_GREEN = colors.HexColor("#00793B")
    ONEE_DARK  = colors.HexColor("#00501F")
    ONEE_LIGHT = colors.HexColor("#E8F5EE")
    title_style = ParagraphStyle("t", fontSize=16, fontName="Helvetica-Bold",
                                  textColor=colors.white)
    sub_style   = ParagraphStyle("s", fontSize=9,  fontName="Helvetica",
                                  textColor=colors.HexColor("#ccddcc"))
    label_style = ParagraphStyle("l", fontSize=11, fontName="Helvetica-Bold",
                                  textColor=ONEE_DARK, spaceAfter=8)
    story = []
    # Header
    ht = Table([[Paragraph(f"ONEE — {title}", title_style),
                 Paragraph(f"Genere le {datetime.now().strftime('%d/%m/%Y %H:%M')}", sub_style)]],
               colWidths=[11*cm, 6*cm])
    ht.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), ONEE_GREEN),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("LEFTPADDING",(0,0),(-1,-1),16),
        ("TOPPADDING",(0,0),(-1,-1),14),
        ("BOTTOMPADDING",(0,0),(-1,-1),14),
    ]))
    story.append(ht)
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(f"Statut : {statut}", label_style))
    # Table
    tdata = [["Parametre","Valeur","Unite"]] + [[p,str(v),u] for p,v,u in data_rows]
    t = Table(tdata, colWidths=[8*cm,5*cm,4*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),   ONEE_DARK),
        ("TEXTCOLOR",     (0,0),(-1,0),   colors.white),
        ("FONTNAME",      (0,0),(-1,0),   "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,-1),  10),
        ("ALIGN",         (0,0),(-1,-1),  "CENTER"),
        ("ALIGN",         (0,1),(0,-1),   "LEFT"),
        ("FONTNAME",      (0,1),(0,-1),   "Helvetica-Bold"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),  [colors.white, ONEE_LIGHT]),
        ("GRID",          (0,0),(-1,-1),  0.5, colors.HexColor("#C8DCC8")),
        ("TOPPADDING",    (0,0),(-1,-1),  8),
        ("BOTTOMPADDING", (0,0),(-1,-1),  8),
        ("LEFTPADDING",   (0,0),(-1,-1),  10),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.8*cm))
    story.append(Paragraph(
        "<font color='#5A7A5A' size='8'>ONEE Tech Assistant v2.1 — Outil interne BT/MT</font>",
        styles["Normal"]))
    doc.build(story)
    buf.seek(0)
    return buf

# ── HEADER ─────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="onee-header">
    <div class="onee-logo">Office National de l'Electricite et de l'Eau Potable</div>
    <div class="onee-title">ONEE Tech Assistant</div>
    <div class="onee-subtitle">Outil de calculs electriques — Reseau Distribution BT/MT</div>
</div>
""", unsafe_allow_html=True)

tab1, tab2 = st.tabs(["🔌 Charge Transformateur", "📉 Chute de Tension"])

# ════════════════════════════════════════════════════════════════════
# TAB 1
# ════════════════════════════════════════════════════════════════════
with tab1:
    st.markdown('<div class="calc-card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">🔌 Calcul de Charge du Transformateur</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        s_nom  = st.number_input("Puissance nominale (kVA)", min_value=1.0, value=100.0, step=10.0)
    with col2:
        p_reel = st.number_input("Puissance active réelle (kW)", min_value=0.1, value=80.0, step=5.0)
    cos_phi_t = st.slider("Facteur de puissance cos φ", 0.50, 1.0, 0.85, 0.01, key="t1")

    if st.button("⚡ Calculer la charge", key="btn_t1"):
        s_reel = p_reel / cos_phi_t
        charge = (s_reel / s_nom) * 100
        q_reel = p_reel * math.tan(math.acos(cos_phi_t))
        if charge > 100:
            box_class, msg = "result-err",  "SURCHARGE — Remplacer ou décharger le transformateur !"
        elif charge > 80:
            box_class, msg = "result-warn", "Charge élevée — Surveillance recommandée."
        else:
            box_class, msg = "result-ok",   "Charge normale — Transformateur dans les limites."
        st.session_state.res_t1 = dict(
            s_nom=s_nom, p_reel=p_reel, cos_phi_t=cos_phi_t,
            s_reel=s_reel, charge=charge, q_reel=q_reel,
            box_class=box_class, msg=msg
        )

    r = st.session_state.res_t1
    if r:
        st.markdown(f"""
        <div class="result-box {r['box_class']}">
            <div class="result-label">Taux de charge</div>
            <div class="result-value">{r['charge']:.1f} %</div>
            <div class="result-msg">{r['msg']}</div>
        </div>""", unsafe_allow_html=True)

        st.markdown("---")
        c1, c2, c3 = st.columns(3)
        c1.metric("S apparente réelle", f"{r['s_reel']:.2f} kVA")
        c2.metric("Q réactive",         f"{r['q_reel']:.2f} kVAR")
        c3.metric("cos φ",              f"{r['cos_phi_t']:.2f}")

        data_rows = [
            ("Puissance nominale",         r['s_nom'],           "kVA"),
            ("Puissance active reelle",    r['p_reel'],          "kW"),
            ("Facteur de puissance",       r['cos_phi_t'],       "—"),
            ("Puissance apparente reelle", round(r['s_reel'],2), "kVA"),
            ("Puissance reactive",         round(r['q_reel'],2), "kVAR"),
            ("Taux de charge",             round(r['charge'],2), "%"),
        ]

        st.markdown("---")
        dl1, dl2 = st.columns(2)
        with dl1:
            st.markdown('<div class="dl-excel">', unsafe_allow_html=True)
            xlsx = make_excel("Rapport Charge Transformateur", data_rows)
            st.download_button("📊 Télécharger Excel", data=xlsx,
                file_name=f"ONEE_Charge_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_excel_t1")
            st.markdown('</div>', unsafe_allow_html=True)
        with dl2:
            st.markdown('<div class="dl-pdf">', unsafe_allow_html=True)
            pdf = make_pdf("Rapport Charge Transformateur", data_rows, r['msg'])
            st.download_button("📄 Télécharger PDF", data=pdf,
                file_name=f"ONEE_Charge_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                mime="application/pdf", key="dl_pdf_t1")
            st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════
# TAB 2
# ════════════════════════════════════════════════════════════════════
with tab2:
    st.markdown('<div class="calc-card">', unsafe_allow_html=True)
    st.markdown('<div class="card-title">📉 Calcul de Chute de Tension — Triphasé</div>', unsafe_allow_html=True)

    st.markdown("**Courants par phase (A)**")
    cp1, cp2, cp3 = st.columns(3)
    with cp1:
        i_ph1 = st.number_input("Phase 1 (A)", min_value=0.0, value=50.0, step=1.0, key="ph1")
    with cp2:
        i_ph2 = st.number_input("Phase 2 (A)", min_value=0.0, value=50.0, step=1.0, key="ph2")
    with cp3:
        i_ph3 = st.number_input("Phase 3 (A)", min_value=0.0, value=50.0, step=1.0, key="ph3")
    i = (i_ph1 + i_ph2 + i_ph3) / 3
    st.info(f"⚡ Courant moyen utilisé pour le calcul : **{i:.2f} A**")

    L = st.number_input("Longueur du câble (m)", min_value=1.0, value=100.0, step=10.0)
    col3, col4 = st.columns(2)
    with col3:
        section  = st.selectbox("Section du câble (mm²)", [16,25,35,50,70,95,120,150])
    with col4:
        material = st.radio("Matériau conducteur", ["Cuivre (Cu)", "Aluminium (Al)"])
    col5, col6 = st.columns(2)
    with col5:
        tension_ref = st.number_input("Tension du réseau (V)", min_value=100.0, value=400.0, step=10.0)
    with col6:
        reseau_type = st.selectbox("Réseau standard", ["Personnalisé", "Triphasé BT (400V)", "Monophasé (230V)", "MT (5500V)"])
        if reseau_type == "Triphasé BT (400V)":
            tension_ref = 400.0
        elif reseau_type == "Monophasé (230V)":
            tension_ref = 230.0
        elif reseau_type == "MT (5500V)":
            tension_ref = 5500.0
    cos_phi_2 = st.slider("Facteur de puissance cos φ", 0.50, 1.0, 0.85, 0.01, key="t2")

    if st.button("⚡ Calculer la chute de tension", key="btn_t2"):
        rho       = 0.0225 if "Cuivre" in material else 0.036
        R         = (rho * L) / section
        delta_u   = math.sqrt(3) * i * R
        perc_drop = (delta_u / tension_ref) * 100
        u_arrive  = tension_ref - delta_u
        s_recommande = None
        if perc_drop > 5:
            box_class, msg = "result-err",  "Chute > 5% — Non conforme NFC 11-201. Augmenter la section !"
            s_min = (rho * L * math.sqrt(3) * i) / (0.05 * tension_ref)
            s_recommande = next((s for s in [16,25,35,50,70,95,120,150,185,240] if s >= s_min), 240)
        elif perc_drop > 3:
            box_class, msg = "result-warn", "Chute entre 3% et 5% — Acceptable mais limite."
        else:
            box_class, msg = "result-ok",   "Chute <= 3% — Conforme aux normes ONEE."
        st.session_state.res_t2 = dict(
            i_ph1=i_ph1, i_ph2=i_ph2, i_ph3=i_ph3, i=i,
            L=L, section=section, material=material,
            cos_phi_2=cos_phi_2, tension_ref=tension_ref, R=R, delta_u=delta_u,
            perc_drop=perc_drop, u_arrive=u_arrive,
            box_class=box_class, msg=msg, s_recommande=s_recommande
        )

    r = st.session_state.res_t2
    if r:
        st.markdown(f"""
        <div class="result-box {r['box_class']}">
            <div class="result-label">Chute de tension</div>
            <div class="result-value">{r['delta_u']:.2f} V &nbsp;|&nbsp; {r['perc_drop']:.2f} %</div>
            <div class="result-msg">{r['msg']}</div>
        </div>""", unsafe_allow_html=True)

        st.markdown("---")
        c1, c2, c3 = st.columns(3)
        c1.metric("Tension départ", f"{r['tension_ref']:.0f} V")
        c2.metric("ΔU calculé",      f"{r['delta_u']:.2f} V")
        c3.metric("Tension arrivée", f"{r['u_arrive']:.1f} V")

        if r['s_recommande']:
            st.info(f"💡 Section minimale recommandée : **{r['s_recommande']} mm²** (pour ΔU ≤ 5%)")

        data_rows = [
            ("Tension réseau",      r['tension_ref'],      "V"),
            ("Courant Phase 1",     r['i_ph1'],            "A"),
            ("Courant Phase 2",     r['i_ph2'],            "A"),
            ("Courant Phase 3",     r['i_ph3'],            "A"),
            ("Courant moyen",       round(r['i'],2),        "A"),
            ("Longueur cable",      r['L'],                "m"),
            ("Section cable",       r['section'],          "mm2"),
            ("Materiau",            r['material'],         "—"),
            ("cos phi",             r['cos_phi_2'],        "—"),
            ("Resistance R",        round(r['R'],4),       "Ohm"),
            ("Chute de tension dU", round(r['delta_u'],2), "V"),
            ("Chute en %",          round(r['perc_drop'],2),"%"),
            ("Tension arrivee",     round(r['u_arrive'],1),"V"),
        ]

        st.markdown("---")
        dl1, dl2 = st.columns(2)
        with dl1:
            st.markdown('<div class="dl-excel">', unsafe_allow_html=True)
            xlsx = make_excel("Rapport Chute de Tension", data_rows)
            st.download_button("📊 Télécharger Excel", data=xlsx,
                file_name=f"ONEE_ChuteTension_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_excel_t2")
            st.markdown('</div>', unsafe_allow_html=True)
        with dl2:
            st.markdown('<div class="dl-pdf">', unsafe_allow_html=True)
            pdf = make_pdf("Rapport Chute de Tension", data_rows, r['msg'])
            st.download_button("📄 Télécharger PDF", data=pdf,
                file_name=f"ONEE_ChuteTension_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                mime="application/pdf", key="dl_pdf_t2")
            st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

# ── Footer ──────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="onee-footer">
    <span class="badge">ONEE Tech v2.1</span> &nbsp;|&nbsp;
    Outil interne de calculs electriques — Distribution BT/MT &nbsp;|&nbsp;
    {datetime.now().strftime('%Y')}
</div>
""", unsafe_allow_html=True)
